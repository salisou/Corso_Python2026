import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

class ModuloBase:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Modulo con classe")
        self.root.geometry("300")
        self.root.configure(bg="#BCC3D1")
        
        self.campi = {}
        
        self.crea_interfaccia()
    
    #-----------------------------------
    #        Placehorder
    #-----------------------------------        
    def set_placehalder(self, entry, testo):
        self.entry = entry
        self.testo = testo
        
        def on_focus_in(event):
            if entry.get() == testo:
                entry.delete(0, tk.END)
                entry.config(fg="black")
                
        def on_focus_out(event):
            if entry.get() == testo:
                entry.inset(0, tk.END)
                entry.config(fg="gry")
    

        entry.bind("<FocusIn>", on_focus_in)
        entry.bind("<FocusOut>", on_focus_out)
        
    #-----------------------------------
    #        Creazione dei campi
    #-----------------------------------
    def crea_campo(self, nome, placeholder, riga, colonna):
        lable = tk.Label(self.root, text = nome, bg= "#e5e7eb")
        lable.grid(row=riga, column=colonna, padx=20, pady=5, sticky="w")
        
        entry = tk.Entry(self.root, bg="#CBD5E1", relief="flat")
        entry.grid(row=riga+1, column=colonna, padx=20, pady=5, ipady=6)
        
        self.set_placehalder(entry, placeholder)
        
        self.campi[nome] = entry
    
    #-----------------------------------
    #        Creazione del interfaccia
    #-----------------------------------
    def crea_interfaccia(self):
        self.crea_campo("Nome", "Inserisci il Nome", 0, 0)
        self.crea_campo("Cognome", "Inserisci il Cognome", 0, 1)
        self.crea_campo("Email", "Inserisci indeizzo mail", 2, 0)
        self.crea_campo("Data", "Inserisci la tua data di nascita", 2, 1)
        
        btn = tk.Button(self.root, text="Salva", command= self.salva_excel)
        btn.grid(row=4, column=0, columnspan=2, pady=20)
        
    #-----------------------------------
    #        Legge Dati
    #-----------------------------------
    def leggi_dati(self):
        valori = []
        for nome, campo in self.campi.items():
            testo = campo.get()
            if campo.cget("fg") == "gray":
                testo = ""
            valori.append(testo)
        return valori
    
    
    #-----------------------------------
    #        Salvataggio Excel
    #  prende i dati
    #  controlla sel il file esiste
    #  se no -> crea
    #  scrive griga
    #-----------------------------------
    def salva_excel(self):
        dati = self.leggi_dati()
        file = "dati.xlsx"
        
        if os.path.exists(file):
            wb = load_workbook(file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(list(self.campi.keys()))
        
        
        ws.append(dati)
        wb.save(file)
        
        messagebox.showinfo("ok", "dati salvati con successo!🎉")    
    #-----------------------------------
    def avvia(self):
        self.root.mainloop()

# Avvio del programma
app = ModuloBase()
app.avvia()